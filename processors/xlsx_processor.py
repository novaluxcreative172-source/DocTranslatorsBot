"""
processors/xlsx_processor.py
Translates a .xlsx workbook cell by cell, across every sheet. Only string
cells are touched -- numbers, dates, and formulas are left exactly as-is
so calculations don't break.
"""

from openpyxl import load_workbook
from translator import translate_text


def process(input_path: str, output_path: str, target_lang: str) -> None:
    wb = load_workbook(input_path)

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip():
                    # Skip formulas (they start with '=') so we don't corrupt them.
                    if cell.value.startswith("="):
                        continue
                    cell.value = translate_text(cell.value, target_lang)

    wb.save(output_path)
